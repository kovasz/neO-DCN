import argparse
import os
import glob
from pathlib import PurePosixPath
from openpyxl import load_workbook
import json

startingRow = 0
startingColumn = "B"

class DcnModel(object):
	def __init__(self, inputFile, outDir):
		self.inputFile = inputFile
		self.outDir = outDir

	def generateBenchmark(self):
		content = self.generateHeader()
		content["links"] = self.generateLinks()
		content["flows"] = self.generateFlows()

		with open("{}/{}.dcn".format(self.outDir, PurePosixPath(self.inputFile).stem), "w") as outFile:
			outFile.write(json.dumps(content, indent = 4))
			outFile.close()
	
	def generateHeader(self):
		raise NotImplementedError()

	def generateLinks(self):
		return [
		{
			"source": 1,
			"destination": 17
		},
		{
			"source": 2,
			"destination": 17
		},
		{
			"source": 3,
			"destination": 18
		},
		{
			"source": 4,
			"destination": 18
		},
		{
			"source": 5,
			"destination": 19
		},
		{
			"source": 6,
			"destination": 19
		},
		{
			"source": 7,
			"destination": 20
		},
		{
			"source": 8,
			"destination": 20
		},
		{
			"source": 9,
			"destination": 21
		},
		{
			"source": 10,
			"destination": 21
		},
		{
			"source": 11,
			"destination": 22
		},
		{
			"source": 12,
			"destination": 22
		},
		{
			"source": 13,
			"destination": 23
		},
		{
			"source": 14,
			"destination": 23
		},
		{
			"source": 15,
			"destination": 24
		},
		{
			"source": 16,
			"destination": 24
		},
		{
			"source": 17,
			"destination": 25
		},
		{
			"source": 17,
			"destination": 26
		},
		{
			"source": 18,
			"destination": 25
		},
		{
			"source": 18,
			"destination": 26
		},
		{
			"source": 19,
			"destination": 27
		},
		{
			"source": 19,
			"destination": 28
		},
		{
			"source": 20,
			"destination": 27
		},
		{
			"source": 20,
			"destination": 28
		},
		{
			"source": 21,
			"destination": 29
		},
		{
			"source": 21,
			"destination": 30
		},
		{
			"source": 22,
			"destination": 29
		},
		{
			"source": 22,
			"destination": 30
		},
		{
			"source": 23,
			"destination": 31
		},
		{
			"source": 23,
			"destination": 32
		},
		{
			"source": 24,
			"destination": 31
		},
		{
			"source": 24,
			"destination": 32
		},
		{
			"source": 25,
			"destination": 33
		},
		{
			"source": 25,
			"destination": 34
		},
		{
			"source": 26,
			"destination": 35
		},
		{
			"source": 26,
			"destination": 36
		},
		{
			"source": 27,
			"destination": 33
		},
		{
			"source": 27,
			"destination": 34
		},
		{
			"source": 28,
			"destination": 35
		},
		{
			"source": 28,
			"destination": 36
		},
		{
			"source": 29,
			"destination": 33
		},
		{
			"source": 29,
			"destination": 34
		},
		{
			"source": 30,
			"destination": 35
		},
		{
			"source": 30,
			"destination": 36
		},
		{
			"source": 31,
			"destination": 33
		},
		{
			"source": 31,
			"destination": 34
		},
		{
			"source": 32,
			"destination": 35
		},
		{
			"source": 33,
			"destination": 36
		}
			]

	def generateFlows(self):
		sheet = load_workbook(self.inputFile).active

		flows = []
		
		print("Generating rows", end = " ")

		i = startingRow
		while i < sheet.max_row and sheet[startingColumn][i].value:
			print(i + 1, end = " ")
			flows.append(self.generateFlow(sheet, startingColumn, i))
			i += 1
		print()
		
		return flows

	def generateFlow(self, sheet, col, row):
		return {
				"source": sheet[col][row].value,
				"destination": sheet[chr(ord(col) + 1)][row].value,
				"packet rate": round(sheet[chr(ord(col) + 2)][row].value)
			}

class DcnModel1(DcnModel):
	def __init__(self, inputFile, outDir):
		super().__init__(inputFile, outDir)

	def generateHeader(self):
		return {
			"version": 1,
			"number of switches": 36
		}

class DcnModel2(DcnModel):
	def __init__(self, inputFile, outDir):
		super().__init__(inputFile, outDir)

	def generateHeader(self):
		return {
			"version": 2,
			"number of servers": 16,
			"number of switches": 20
		}


parser = argparse.ArgumentParser("Convert an Excel file (used as input for LINGO) to a JSON file (as input for neO)")

parser.add_argument("path", help="path to XLSX file(s)")
parser.add_argument("--model",
					action="store", type=int, dest="model", default=1,
					help="version numbder of the DCN model")

args = parser.parse_args()

model = args.model - 1
dcnModels = [DcnModel1,DcnModel2]

# Checking and collecting the input files
if not os.path.exists(args.path):
	print("Input path {} does not exist".format(args.path))
	exit()
inputFiles = []
if os.path.isdir(args.path):
	for file in glob.glob("{}/*.xlsx".format(args.path)):
		inputFiles.append(file)
else:
	inputFiles.append(args.path)

# Creating and cleaning out directory
outDir = "./out"
if not os.path.exists(outDir):
	os.makedirs(outDir)

for file in inputFiles:
	print("Converting {}".format(file))
	dcnModel = dcnModels[model](file, outDir)
	dcnModel.generateBenchmark()

print("Files converted!")